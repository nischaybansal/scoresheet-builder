"""
Scoresheet generator engine.
Given a competency framework (list of (competency, behaviour)) and a list of
selected tools, builds the multi-tab assessor scoresheet workbook:
  - one sheet per selected ONLINE tool (OPQ/MQ/SJT) with the paste-stens engine
  - one sheet per selected OFFLINE tool (Case Study/Role Play/BEI) as a blank scaffold
  - a Detailed Integration sheet wiring every tool score into place
Returns an openpyxl Workbook.
"""
from itertools import groupby
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

# ---- constant SHL dimension libraries (captured from the sample workbook) ----
OPQ_DIMS = ["Persuasive","Controlling","Outspoken","Independent Minded","Outgoing","Affiliative",
"Socially Confident","Modest","Democratic","Caring","Data Rational","Evaluative","Behavioural",
"Conventional","Conceptual","Innovative","Variety Seeking","Adaptable","Forward Thinking",
"Detail Conscious","Conscientious","Rule Following","Relaxed","Worrying","Tough Minded","Optimistic",
"Trusting","Emotionally Controlled","Vigorous","Competitive","Achieving","Decisive"]
MQ_DIMS = ["Level of Activity","Achievement","Competition","Fear of Failure","Power","Immersion",
"Commercial Outlook","Affiliation","Recognition","Personal Principles","Ease and Security",
"Personal Growth","Interest","Flexibility","Autonomy","Material Reward","Progression","Status"]
SJT_VARIANTS = {
    "Managerial": {"group":"Scenarios Management Profile",
        "dims":["Personal Recognition","Big Picture","Company Protocol","Delegative",
                "Managerial Judgement","Managing Objectives","One To One","People Management",
                "Reputation Management","Team"]},
    "Executive": {"group":"Scenarios Executive Profile",
        "dims":["Corporate Management","Managerial Judgement","Managing Objectives","People Management"]},
}
DEFAULT_SJT_VARIANT = "Managerial"

# ---- tool registry: canonical order = online first, then offline ----
# kind: 'online' (paste-stens), 'bars' (rubric anchors), 'bei' (interview questions)
# score_col = the column on the tool sheet holding the behaviour's 1-5 rating that DI reads
VERIFY_DIMS = ["Inductive Score","Deductive Score","Numerical Score"]

# ---- Onlines Data sheet layout (mirrors the real SHL export) ----
OD_SHEET = "Onlines Data"
OD_IDENTITY = ["Project Name", "First Name", "Last Name", "Email", "Completed Date", "Status"]
OD_EMAIL_COL = "D"             # Email is the 4th identity column
OD_FIRST_COL = "B"             # First Name
OD_LAST_COL  = "C"             # Last Name
OD_HELPER_ROW = 4              # hidden row: one MATCH per column (keeps the workbook fast)
OD_FIRST_DATA_ROW = 5          # participant rows auto-pulled from the Consultant Paste sheet
OD_LAST_DATA_ROW  = 504        # 500 participants headroom (dynamic - blanks ignored)
OD_TRAIT_FIRST_COL = "G"       # traits start after the six identity columns
OD_TRAIT_LAST_COL  = "DZ"      # headroom for extra export columns
EMAIL_REF = "'Detailed Integration'!$B$2"   # selected participant email
CONSULT_SHEET = "Consultant Paste"
CP_GROUP_ROW  = 2              # the export's group row (MQ.M5 / OPQ32r / Scenarios ...) —
                               # identity names land here because the export merges them
                               # vertically down into the trait row
CP_HEADER_ROW = 3              # the export's trait header row
CP_FIRST_DATA_ROW = 4          # first participant
CP_LAST_COL = "DZ"

# Raw export header text (Onlines Data row 2). Lookups match on this FIRST and fall back to
# the clean trait name in row 3, so either form in the pasted header row works.
_MQ_CODES = ["E1", "E2", "E3", "E4", "E5", "E6", "E7", "S1", "S2", "S3", "S4", "S5",
             "I1", "I2", "I3", "X1", "X2", "X3"]
_OPQ_CODES = ["RP1", "RP2", "RP3", "RP4", "RP5", "RP6", "RP7", "RP8", "RP9", "RP10",
              "TS1", "TS2", "TS3", "TS4", "TS5", "TS6", "TS7", "TS8", "TS9", "TS10",
              "TS11", "TS12", "FE1", "FE2", "FE3", "FE4", "FE5", "FE6", "FE7", "FE8",
              "FE9", "FE10"]


def raw_header(trait):
    """The exact header text the SHL export produces for a trait."""
    # NOTE: check trait lists before identity — "Status" is both an identity column
    # and an MQ dimension (Status X3), and the trait meaning must win here.
    if trait in MQ_DIMS:
        return f"{trait} ({_MQ_CODES[MQ_DIMS.index(trait)]})-Sten"
    if trait in OPQ_DIMS:
        return f"{trait} ({_OPQ_CODES[OPQ_DIMS.index(trait)]})-Sten"
    if trait == "Consistency Measure":
        return "Consistency Measure (CNS)-STEN"
    for v in SJT_VARIANTS.values():
        if trait in v["dims"]:
            return f"{trait}-STEN"
    if trait in OD_IDENTITY:
        return trait
    return trait          # Verify and anything else: header text not yet confirmed


TOOLS = {
    "OPQ":              {"kind":"online","dims":OPQ_DIMS,   "family":"online"},
    "MQ":               {"kind":"online","dims":MQ_DIMS,    "family":"online"},
    "SJT":              {"kind":"online","dims":None,       "family":"online"},  # dims from SJT_VARIANTS
    "Verify":           {"kind":"online","dims":VERIFY_DIMS,"family":"online"},
    "Case Study":       {"kind":"bars","score_col":"H","family":"offline"},
    "Group Discussion": {"kind":"bars","score_col":"H","family":"offline"},
    "Inbox Simulation": {"kind":"bars","score_col":"H","family":"offline"},
    "Written Analysis": {"kind":"bars","score_col":"H","family":"offline"},
    "Business Role Play": {"kind":"bars","score_col":"H","family":"offline"},
    "Coaching Role Play": {"kind":"bars","score_col":"H","family":"offline"},
    "BEI":              {"kind":"bei","score_col":"F","family":"offline"},
}
CANONICAL = ["OPQ","MQ","SJT","Verify",
             "Case Study","Group Discussion","Inbox Simulation","Written Analysis",
             "Business Role Play","Coaching Role Play","BEI"]
def _meta_cols(meta):
    """(has_theme, has_level, theme_label, level_label) for whatever the client supplied."""
    meta = meta or {}
    return (bool(meta.get("themes")), bool(meta.get("levels")),
            meta.get("theme_label") or "Theme", meta.get("level_label") or "Level")


def _score_col(name, meta=None):
    """Where a tool's rating lands, once optional Theme/Level columns are counted."""
    has_theme, has_level, _, _ = _meta_cols(meta)
    extra = int(has_theme) + int(has_level)
    base = {"bars": 8, "bei": 6}.get(TOOLS[name]["kind"])
    if base is None:
        return TOOLS[name].get("score_col")
    return openpyxl.utils.get_column_letter(base + extra)


def _avg_col(name, mq_mode="5-point", meta=None):
    """Where the Competency Avg column lands on an online sheet. Computed rather than fixed,
    because optional Theme/Level columns and OPQ's Direction column all shift it."""
    has_theme, has_level, _, _ = _meta_cols(meta)
    n_pre = 2 + int(has_theme) + int(has_level)      # [theme] competency [level] behaviour
    nrub = 2 if (name == "MQ" and mq_mode == "low-high") else 5
    has_dir = 1 if name == "OPQ" else 0
    idx = n_pre + 1 + has_dir + nrub + 3             # trait, [direction], rubrics, sten/score/avg
    return openpyxl.utils.get_column_letter(idx)

HR = 6          # header row on offline sheets
FB = HR + 1     # first behaviour row on offline sheets
# online sheets carry the sten block at the top, so their table starts lower
OB_NAME_ROW = 3          # trait names run across this row
OB_STEN_ROW = 4          # stens run across the row beneath
ONLINE_HR = 9            # header row on online sheets
ONLINE_FB = ONLINE_HR + 1

# ---- styling ----
ARIAL="Montserrat"   # all cells use Montserrat 9 (must be installed for Excel to render it)
C_HDR=PatternFill("solid",fgColor="1F3864"); C_SUB=PatternFill("solid",fgColor="D9E1F2")
C_IN=PatternFill("solid",fgColor="FFF2CC");  C_LOCK=PatternFill("solid",fgColor="E2EFDA")
C_LEG=PatternFill("solid",fgColor="FCE4D6"); C_PASTE=PatternFill("solid",fgColor="DDEBF7")
C_DIMBG=PatternFill("solid",fgColor="F2F2F2"); C_BLACK=PatternFill("solid",fgColor="000000")
C_BAD=PatternFill("solid",fgColor="FFC7CE")    # a real conflict
C_WARN=PatternFill("solid",fgColor="FFEB9C")   # not filled in yet
C_NEG=PatternFill("solid",fgColor="FFD6D6")    # reverse-scored (Negative) trait row - light
                                               # enough that the text stays readable
F_HDR=Font(name=ARIAL,size=9,bold=True,color="FFFFFF"); F_B=Font(name=ARIAL,size=9,bold=True)
F_N=Font(name=ARIAL,size=9); F_TITLE=Font(name=ARIAL,size=9,bold=True,color="1F3864")
F_IT=Font(name=ARIAL,size=9,italic=True); F_TAG=Font(name=ARIAL,size=9,bold=True,color="1F3864")
_thin=Side(style="thin",color="BFBFBF"); BD=Border(_thin,_thin,_thin,_thin)
A_WRAP=Alignment(wrap_text=True,vertical="top"); A_CTR=Alignment(horizontal="center",vertical="center")
A_WC=Alignment(wrap_text=True,vertical="center",horizontal="center")
A_WV=Alignment(wrap_text=True,vertical="center")

def _ref(sheet):
    return f"'{sheet}'" if " " in sheet else sheet

def _hdr(ws,row,cols):
    for c in cols:
        x=ws[f"{c}{row}"]; x.font=F_HDR; x.fill=C_HDR; x.alignment=A_WC; x.border=BD

def _title(ws,text,legend,span):
    ws["A1"]=text; ws["A1"].font=F_TITLE
    ws["A3"]=legend; ws["A3"].font=F_N; ws["A3"].fill=C_LEG
    ws.merge_cells(f"A3:{span}3"); ws["A3"].alignment=A_WV

def _cellsetup(ws,row,cols):
    for c in cols:
        ws[f"{c}{row}"].border=BD; ws[f"{c}{row}"].font=F_N; ws[f"{c}{row}"].alignment=A_WRAP

def _merge_down(ws,col,gs,ge,centre=True):
    """Merge one column down a competency group so the label is written once, not repeated
    on every behaviour row. Keeps the border on every cell so the grid still reads."""
    if not col or ge<=gs:
        return
    ws.merge_cells(f"{col}{gs}:{col}{ge}")
    ws[f"{col}{gs}"].alignment=(A_WC if centre else A_WV)
    for rr in range(gs,ge+1):
        ws[f"{col}{rr}"].border=BD

def _merge_if_uniform(ws,col,gs,ge,values):
    """Merge a column down a group ONLY when it carries a single value across that group.

    This is what separates the two ways the third framework column gets used: a competency
    DEFINITION is one thing spanning every behaviour, so it merges into one cell; a genuine
    LEVEL that varies behaviour-to-behaviour (Core / Advanced / Expert) stays per row so the
    variation is still visible. Returns True if it merged."""
    if not col or ge<=gs:
        return False
    seen={(v or "").strip() for v in values}
    if len(seen)>1:
        return False
    _merge_down(ws,col,gs,ge)
    return True

# ---------- ONLINE tool sheet (paste-stens engine) ----------
TRAIT_SLOTS = 12  # max mapped traits per competency on OPQ (other online tools cap at their
                  # own dimension count, so raising this only affects OPQ)

def _blocks(framework):
    """[(comp, [behaviours])] preserving order (framework arrives clustered)."""
    out=[]
    for comp,beh in framework:
        if out and out[-1][0]==comp: out[-1][1].append(beh)
        else: out.append((comp,[beh]))
    return out

def _tool_slots(name,dims=None,sjt_variant=DEFAULT_SJT_VARIANT):
    """How many trait slots a competency block gets on an online sheet. OPQ gets the full
    TRAIT_SLOTS; every other online tool caps at its own dimension count. Used by BOTH the
    tool sheet and Detailed Integration, so their row geometry can never drift apart."""
    if name=="OPQ":
        return TRAIT_SLOTS
    if dims is None:
        dims=SJT_VARIANTS[sjt_variant]["dims"] if name=="SJT" else (TOOLS[name].get("dims") or [])
    return min(TRAIT_SLOTS,len(dims))

def _block_geometry(framework,slots=TRAIT_SLOTS):
    """Row layout for an online sheet: block height = max(n_behaviours, slots).
    `slots` is the tool's own trait capacity - OPQ carries the full TRAIT_SLOTS, while the
    other online tools cap at their dimension count, so Verify (3 dims) does not get a block
    of 12 mostly-empty rows. Returns [(comp, behaviours, start_row, end_row)]."""
    geo=[]; r=ONLINE_FB
    for comp,behs in _blocks(framework):
        h=max(len(behs),slots)
        geo.append((comp,behs,r,r+h-1)); r+=h
    return geo

def _sort_block(ws,first_col_idx,blocks,rub_cols,score_col,label,mode="5-point",sten_col=None):
    """Adds the (hidden) statement-selection + descending-sort columns used by the report:
      Rank Key | Statement Selected | Sorted Score | Statements in Order
    blocks: [(start_row,end_row,n_slots)] - one per competency.
    mode "5-point": statement picked by Score 1-5 from the five rubric columns.
    mode "low-high": statement picked only when Sten<=3 (low rubric) or Sten>=8 (high rubric);
    mid stens produce no statement and drop out of the sort."""
    L=openpyxl.utils.get_column_letter
    key_c=L(first_col_idx); sel_c=L(first_col_idx+1); srt_c=L(first_col_idx+2); ord_c=L(first_col_idx+3)
    for c,h in zip([key_c,sel_c,srt_c,ord_c],
                   ["Rank Key",f"{label} Statement Selected","Sorted Score","Statements in Order"]):
        ws[f"{c}{HR}"]=h
    _hdr(ws,HR,[key_c,sel_c,srt_c,ord_c])
    for gs,ge,n in blocks:
        last=gs+n-1
        for i in range(n):
            r=gs+i
            if mode=="low-high":
                lo,hi=rub_cols[0],rub_cols[1]
                ws[f"{sel_c}{r}"]=(f'=IF({sten_col}{r}="","",IF({sten_col}{r}<=3,{lo}{r},'
                                   f'IF({sten_col}{r}>=8,{hi}{r},"")))')
            else:
                e=rub_cols
                ws[f"{sel_c}{r}"]=(f'=IF({score_col}{r}="","",IF({score_col}{r}=1,{e[0]}{r},'
                    f'IF({score_col}{r}=2,{e[1]}{r},IF({score_col}{r}=3,{e[2]}{r},'
                    f'IF({score_col}{r}=4,{e[3]}{r},{e[4]}{r})))))')
            # rank key: score + a tiny row offset so equal scores keep a stable order;
            # rows with no statement drop out of the ranking entirely
            ws[f"{key_c}{r}"]=(f'=IF(OR({score_col}{r}="",{sel_c}{r}=""),"",'
                               f'{score_col}{r}+ROW()*0.0000000001)')
            ws[f"{srt_c}{r}"]=f'=IFERROR(LARGE({key_c}${gs}:{key_c}${last},{i+1}),"")'
            ws[f"{ord_c}{r}"]=(f'=IFERROR(INDEX({sel_c}${gs}:{sel_c}${last},'
                               f'MATCH(LARGE({key_c}${gs}:{key_c}${last},{i+1}),'
                               f'{key_c}${gs}:{key_c}${last},0)),"")')
            for c in [key_c,sel_c,srt_c,ord_c]:
                ws[f"{c}{r}"].font=F_N; ws[f"{c}{r}"].border=BD; ws[f"{c}{r}"].fill=C_LOCK
                ws[f"{c}{r}"].alignment=A_WRAP
    for c,w in zip([key_c,sel_c,srt_c,ord_c],[10,40,10,40]):
        ws.column_dimensions[c].width=w
        ws.column_dimensions[c].hidden=True     # hidden - they feed the report
    return ord_c

def _online_block(ws,name,dims,online_paste):
    """The sten block that runs ACROSS the top of every online sheet: trait names on one
    row, stens on the row beneath. Every mapped trait then reads its sten from here with
    HLOOKUP, so a trait used by three competencies is only ever entered once.
    Returns (lookup_range, names_range)."""
    L=openpyxl.utils.get_column_letter
    od_data=f"'{OD_SHEET}'!${OD_TRAIT_FIRST_COL}${OD_FIRST_DATA_ROW}:${OD_TRAIT_LAST_COL}${OD_LAST_DATA_ROW}"
    od_names=f"'{OD_SHEET}'!${OD_TRAIT_FIRST_COL}$3:${OD_TRAIT_LAST_COL}$3"
    od_email=f"'{OD_SHEET}'!${OD_EMAIL_COL}${OD_FIRST_DATA_ROW}:${OD_EMAIL_COL}${OD_LAST_DATA_ROW}"

    ws[f"A{OB_NAME_ROW}"]=f"{name} Dimensions"
    ws[f"A{OB_STEN_ROW}"]=("Sten (auto)" if online_paste else "Paste Sten here")
    for cc in (OB_NAME_ROW,OB_STEN_ROW):
        c=ws[f"A{cc}"]; c.font=F_HDR; c.fill=C_HDR; c.alignment=A_WC; c.border=BD

    for i,dim in enumerate(dims):
        col=L(2+i)
        n=ws[f"{col}{OB_NAME_ROW}"]
        n.value=dim; n.font=F_HDR; n.fill=C_HDR; n.alignment=A_WC; n.border=BD
        s=ws[f"{col}{OB_STEN_ROW}"]
        if online_paste:
            _ix=(f'INDEX({od_data},MATCH({EMAIL_REF},{od_email},0),'
                 f'MATCH({col}${OB_NAME_ROW},{od_names},0))')
            s.value=f'=IF({EMAIL_REF}="","",IFERROR(IF({_ix}="","",{_ix}),""))'
            s.fill=C_LOCK
        else:
            s.value=""; s.fill=C_PASTE
        s.font=F_N; s.alignment=A_CTR; s.border=BD
    last=L(1+len(dims))
    return (f"$B${OB_NAME_ROW}:${last}${OB_STEN_ROW}", f"$B${OB_NAME_ROW}:${last}${OB_NAME_ROW}")


def _build_online(wb,name,dims,framework,mq_mode="5-point",online_paste=True,meta=None):
    from openpyxl.styles import Protection
    has_dir = (name=="OPQ")
    low_high = (name=="MQ" and mq_mode=="low-high")
    slots = TRAIT_SLOTS if name=="OPQ" else min(TRAIT_SLOTS,len(dims))
    ws=wb.create_sheet(name)
    disp=name
    key=("SCORING KEY (constant): Positive trait -> Score = roundup(Sten/2): 1-2=1 · 3-4=2 · 5-6=3 · 7-8=4 · 9-10=5.  "
         "Negative trait -> inverse: 1-2=5 · 3-4=4 · 5-6=3 · 7-8=2 · 9-10=1.") if has_dir else \
        ("SCORING KEY (constant): Score = roundup(Sten/2): 1-2=1 · 3-4=2 · 5-6=3 · 7-8=4 · 9-10=5.")
    avg=_avg_col(name,mq_mode,meta)
    if name=="SJT":
        # label which variant this sheet was generated for
        disp=f"SJT ({[k for k,v in SJT_VARIANTS.items() if v['dims']==dims][0]})" if any(v['dims']==dims for v in SJT_VARIANTS.values()) else name
    ws["A1"]=f"{disp}  —  online psychometric"; ws["A1"].font=F_TITLE
    blk,blk_names=_online_block(ws,name,dims,online_paste)
    ws["A6"]=key; ws["A6"].font=F_N; ws["A6"].fill=C_LEG
    ws.merge_cells(f"A6:{avg}6"); ws["A6"].alignment=A_WV
    ws["A7"]=(("Stens fill themselves from the 'Onlines Data' sheet for the participant "
               "selected in Detailed Integration. " if online_paste else
               "Paste every sten into the block at the top - each trait once. ")
              + "Consultant maps up to "
              f"{slots} traits per competency and writes the rubric descriptions"
              +(" and sets Direction (Positive by default)" if has_dir else "")
              +(". This sheet is protected - only mapping"
                +(", direction" if has_dir else "")+" and rubric cells are editable."
                if online_paste else "."))
    ws["A7"].font=F_IT; ws.merge_cells(f"A7:{avg}7"); ws["A7"].alignment=A_WV
    rub_heads=(["Low Rubric (Sten 1-3)","High Rubric (Sten 8-10)"] if low_high
               else ["Rating 1","Rating 2","Rating 3","Rating 4","Rating 5"])
    has_theme,has_level,theme_lbl,level_lbl=_meta_cols(meta)
    pre=([theme_lbl] if has_theme else [])+["Competency"]+([level_lbl] if has_level else [])+["Behaviour"]
    heads=pre+[f"Mapped {name} Trait"]+(["Direction"] if has_dir else [])+ \
          rub_heads+["Sten (auto)","Score (1-5)","Competency Avg"]
    cols=[openpyxl.utils.get_column_letter(i+1) for i in range(len(heads))]
    theme_col=cols[0] if has_theme else None
    comp_col=cols[1] if has_theme else cols[0]
    level_col=cols[len(pre)-2] if has_level else None
    beh_col_o=cols[len(pre)-1]
    trait_col=cols[len(pre)]
    dir_col=cols[len(pre)+1] if has_dir else None
    nrub=len(rub_heads)
    rub_start=len(pre)+(2 if has_dir else 1)
    rub_cols=cols[rub_start:rub_start+nrub]
    sten_col=cols[-3]; score_col=cols[-2]
    for c,h in zip(cols,heads): ws[f"{c}{ONLINE_HR}"]=h
    _hdr(ws,ONLINE_HR,cols)
    unlocked=Protection(locked=False)
    geo=_block_geometry(framework,slots)
    for comp,behs,gs,ge in geo:
        lvl_vals=[]
        for k in range(ge-gs+1):
            r=gs+k
            if k<len(behs):
                ws[f"{comp_col}{r}"]=comp; ws[f"{beh_col_o}{r}"]=behs[k]
                if has_theme:
                    ws[f"{theme_col}{r}"]=(meta.get("themes") or {}).get(comp,"")
                if has_level:
                    lv=(meta.get("levels") or {}).get((comp,behs[k]),"")
                    ws[f"{level_col}{r}"]=lv; lvl_vals.append(lv)
            if k<slots:
                ws[f"{trait_col}{r}"]=""; ws[f"{trait_col}{r}"].fill=C_IN
                if has_dir:
                    ws[f"{dir_col}{r}"]="Positive"; ws[f"{dir_col}{r}"].fill=C_IN
                for rc in rub_cols:
                    ws[f"{rc}{r}"]=""; ws[f"{rc}{r}"].fill=C_SUB
                ws[f"{sten_col}{r}"]=(f'=IF({trait_col}{r}="","",'
                    f'IFERROR(HLOOKUP({trait_col}{r},{blk},2,FALSE),""))')
                ws[f"{sten_col}{r}"].fill=C_LOCK
                if has_dir:
                    ws[f"{score_col}{r}"]=(f'=IF({sten_col}{r}="","",IF({dir_col}{r}="Negative",'
                        f'6-ROUNDUP({sten_col}{r}/2,0),ROUNDUP({sten_col}{r}/2,0)))')
                else:
                    ws[f"{score_col}{r}"]=f'=IF({sten_col}{r}="","",ROUNDUP({sten_col}{r}/2,0))'
                ws[f"{score_col}{r}"].fill=C_LOCK
                editable=[trait_col]+([dir_col] if has_dir else [])+rub_cols
                for cc in editable:
                    ws[f"{cc}{r}"].protection=unlocked
            _cellsetup(ws,r,cols)
            ctr_cols=[sten_col,score_col]+([dir_col] if has_dir else [])
            for cc in ctr_cols: ws[f"{cc}{r}"].alignment=A_CTR
        g=ws[f"{avg}{gs}"]
        g.value=f'=IFERROR(ROUND(AVERAGE({score_col}{gs}:{score_col}{ge}),0),"")'
        g.fill=C_LOCK; g.font=F_B; g.alignment=A_CTR
        if ge>gs: ws.merge_cells(f"{avg}{gs}:{avg}{ge}")
        for rr in range(gs,ge+1): ws[f"{avg}{rr}"].border=BD
        # theme / competency written once per block; the third column merges only when it
        # holds one value across the competency (a definition), not when it varies (a level)
        if has_theme: _merge_down(ws,theme_col,gs,ge)
        _merge_down(ws,comp_col,gs,ge)
        if has_level: _merge_if_uniform(ws,level_col,gs,ge,lvl_vals)
    last=geo[-1][3]
    dv=DataValidation(type="list",formula1=f"={blk_names}",allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"{trait_col}{ONLINE_FB}:{trait_col}{last}")
    if has_dir:
        dv2=DataValidation(type="list",formula1='"Positive,Negative"',allow_blank=True)
        ws.add_data_validation(dv2); dv2.add(f"{dir_col}{ONLINE_FB}:{dir_col}{last}")
        # reverse-scored traits (Negative inverts to 6-ROUNDUP(sten/2)) are flagged by
        # tinting the WHOLE trait row light red - dark enough to spot, light enough to read
        ws.conditional_formatting.add(
            f"{cols[0]}{ONLINE_FB}:{cols[-1]}{last}",
            FormulaRule(formula=[f'${dir_col}{ONLINE_FB}="Negative"'],
                        fill=C_NEG,stopIfTrue=False))
    widths=([18] if has_theme else [])+[20]+([12] if has_level else [])+[38,22] \
           +([10] if has_dir else [])+[17]*nrub+[10,9,12]
    for c,w in zip(cols,widths): ws.column_dimensions[c].width=w
    ws.freeze_panes=f"A{ONLINE_FB}"
    # hidden statement-selection + descending-sort columns (feed the report)
    blocks=[(gs,ge,min(slots,ge-gs+1)) for comp,behs,gs,ge in geo]
    _sort_block(ws,len(cols)+5,blocks,rub_cols,score_col,name,
                mode=("low-high" if low_high else "5-point"),sten_col=sten_col)
    ws.protection.sheet=True     # lock the sheet; only unlocked cells editable

def _build_mq_standalone(wb, dims, online_paste=True):
    """MQ in Low/High mode is a standalone motivational profile: it is NOT mapped to any
    competency and does NOT appear in Detailed Integration or Integration. One row per
    trait, low/high rubrics, auto sten and score, plus the hidden sorted-statement columns."""
    from openpyxl.styles import Protection
    ws = wb.create_sheet("MQ")
    ws["A1"] = "MQ  —  motivational profile (Low / High)"; ws["A1"].font = F_TITLE
    blk, blk_names = _online_block(ws, "MQ", dims, online_paste)
    ws["A6"] = ("SCORING KEY (constant): Score = roundup(Sten/2): 1-2=1 · 3-4=2 · 5-6=3 · 7-8=4 · 9-10=5.  "
                "A statement is produced only for Sten 1-3 (low rubric) or Sten 8-10 (high rubric).")
    ws["A6"].font = F_N; ws["A6"].fill = C_LEG
    ws.merge_cells("A6:E6"); ws["A6"].alignment = A_WV
    ws["A7"] = ("Not mapped to competencies - this profile stands on its own and does not feed "
                "Detailed Integration. " + ("Stens fill themselves from the participant selected "
                "in Detailed Integration. This sheet is protected; only the rubric cells are "
                "editable." if online_paste else "Paste every sten into the block at the top."))
    ws["A7"].font = F_IT; ws.merge_cells("A7:E7"); ws["A7"].alignment = A_WV

    heads = ["MQ Trait", "Low Rubric (Sten 1-3)", "High Rubric (Sten 8-10)",
             "Sten (auto)", "Score (1-5)"]
    cols = ["A", "B", "C", "D", "E"]
    for c, h in zip(cols, heads):
        ws[f"{c}{ONLINE_HR}"] = h
    _hdr(ws, ONLINE_HR, cols)
    unlocked = Protection(locked=False)

    for i, dim in enumerate(dims):
        r = ONLINE_FB + i
        ws[f"A{r}"] = dim; ws[f"A{r}"].fill = C_DIMBG
        for c in ("B", "C"):
            ws[f"{c}{r}"] = ""; ws[f"{c}{r}"].fill = C_SUB
            ws[f"{c}{r}"].protection = unlocked
        ws[f"D{r}"] = f'=IFERROR(HLOOKUP(A{r},{blk},2,FALSE),"")'
        ws[f"D{r}"].fill = C_LOCK
        ws[f"E{r}"] = f'=IF(D{r}="","",ROUNDUP(D{r}/2,0))'
        ws[f"E{r}"].fill = C_LOCK
        _cellsetup(ws, r, cols)
        for c in ("D", "E"):
            ws[f"{c}{r}"].alignment = A_CTR

    last = ONLINE_FB + len(dims) - 1
    # one sorted block across the whole profile (no competency grouping)
    _sort_block(ws, len(cols) + 3, [(ONLINE_FB, last, len(dims))], ["B", "C"], "E", "MQ",
                mode="low-high", sten_col="D")
    for c, w in zip(cols, [26, 34, 34, 10, 10]):
        ws.column_dimensions[c].width = w
    ws.freeze_panes = f"A{ONLINE_FB}"
    ws.protection.sheet = True
    return ws


# ---------- OFFLINE: BARS rubric sheet (Case Study / Role Play) ----------
def _build_bars(wb,name,framework,meta=None):
    ws=wb.create_sheet(name)
    has_theme,has_level,theme_lbl,level_lbl=_meta_cols(meta)
    pre=([theme_lbl] if has_theme else [])+["Competency"]+([level_lbl] if has_level else [])+["Behaviour"]
    heads=pre+["Rating 1","Rating 2","Rating 3","Rating 4","Rating 5","Rating","Comments"]
    cols=[openpyxl.utils.get_column_letter(i+1) for i in range(len(heads))]
    _title(ws,f"{name}  —  scoring guidelines",
        "Consultant writes the anchors under Rating 1-5. Assessor enters observed Rating (1-5) + comments.",
        cols[-1])
    theme_col=cols[0] if has_theme else None
    comp_col=cols[1] if has_theme else cols[0]
    level_col=cols[len(pre)-2] if has_level else None
    beh_col=cols[len(pre)-1]
    rub=cols[len(pre):len(pre)+5]
    rate_col=cols[len(pre)+5]; cmt_col=cols[len(pre)+6]
    for c,h in zip(cols,heads): ws[f"{c}{HR}"]=h
    _hdr(ws,HR,cols)
    for i,(comp,beh) in enumerate(framework):
        r=FB+i
        ws[f"{comp_col}{r}"]=comp; ws[f"{beh_col}{r}"]=beh
        if has_theme: ws[f"{theme_col}{r}"]=(meta.get("themes") or {}).get(comp,"")
        if has_level: ws[f"{level_col}{r}"]=(meta.get("levels") or {}).get((comp,beh),"")
        for c in rub: ws[f"{c}{r}"]=""; ws[f"{c}{r}"].fill=C_SUB
        ws[f"{rate_col}{r}"]=""; ws[f"{rate_col}{r}"].fill=C_IN
        ws[f"{cmt_col}{r}"]=""; ws[f"{cmt_col}{r}"].fill=C_IN
        _cellsetup(ws,r,cols); ws[f"{rate_col}{r}"].alignment=A_CTR
    # write theme / competency once per group; level merges only if uniform
    r=FB
    for comp,behs in _blocks(framework):
        gs=r; ge=r+len(behs)-1
        if has_theme: _merge_down(ws,theme_col,gs,ge)
        _merge_down(ws,comp_col,gs,ge)
        if has_level:
            _merge_if_uniform(ws,level_col,gs,ge,
                              [(meta.get("levels") or {}).get((comp,b),"") for b in behs])
        r=ge+1
    widths=([18] if has_theme else [])+[18]+([12] if has_level else [])+[36,18,18,18,18,18,8,28]
    for c,w in zip(cols,widths): ws.column_dimensions[c].width=w
    ws.freeze_panes=f"A{FB}"
    # hidden statement-selection + descending-sort columns (feed the report)
    blocks=[]; r=FB
    for comp,behs in _blocks(framework):
        blocks.append((r,r+len(behs)-1,len(behs))); r+=len(behs)
    if blocks:
        _sort_block(ws,len(cols)+3,blocks,rub,rate_col,name)


def _build_bei(wb,name,framework,meta=None):
    ws=wb.create_sheet(name)
    has_theme,has_level,theme_lbl,level_lbl=_meta_cols(meta)
    pre=([theme_lbl] if has_theme else [])+["Competency"]+([level_lbl] if has_level else [])+["Behaviour"]
    heads=pre+["BEI Question 1","BEI Question 2","BEI Question 3","Rating","Comments"]
    cols=[openpyxl.utils.get_column_letter(i+1) for i in range(len(heads))]
    _title(ws,f"{name}  —  Behavioural Event Interview",
        "Consultant writes 1-3 BEI questions per behaviour. Assessor enters Rating (1-5) + comments.",
        cols[-1])
    theme_col=cols[0] if has_theme else None
    comp_col=cols[1] if has_theme else cols[0]
    level_col=cols[len(pre)-2] if has_level else None
    beh_col=cols[len(pre)-1]
    qs=cols[len(pre):len(pre)+3]
    rate_col=cols[len(pre)+3]; cmt_col=cols[len(pre)+4]
    for c,h in zip(cols,heads): ws[f"{c}{HR}"]=h
    _hdr(ws,HR,cols)
    for i,(comp,beh) in enumerate(framework):
        r=FB+i
        ws[f"{comp_col}{r}"]=comp; ws[f"{beh_col}{r}"]=beh
        if has_theme: ws[f"{theme_col}{r}"]=(meta.get("themes") or {}).get(comp,"")
        if has_level: ws[f"{level_col}{r}"]=(meta.get("levels") or {}).get((comp,beh),"")
        for c in qs: ws[f"{c}{r}"]=""; ws[f"{c}{r}"].fill=C_SUB
        ws[f"{rate_col}{r}"]=""; ws[f"{rate_col}{r}"].fill=C_IN
        ws[f"{cmt_col}{r}"]=""; ws[f"{cmt_col}{r}"].fill=C_IN
        _cellsetup(ws,r,cols); ws[f"{rate_col}{r}"].alignment=A_CTR
    # write theme / competency once per group; level merges only if uniform
    r=FB
    for comp,behs in _blocks(framework):
        gs=r; ge=r+len(behs)-1
        if has_theme: _merge_down(ws,theme_col,gs,ge)
        _merge_down(ws,comp_col,gs,ge)
        if has_level:
            _merge_if_uniform(ws,level_col,gs,ge,
                              [(meta.get("levels") or {}).get((comp,b),"") for b in behs])
        r=ge+1
    widths=([18] if has_theme else [])+[18]+([12] if has_level else [])+[36,26,26,26,8,28]
    for c,w in zip(cols,widths): ws.column_dimensions[c].width=w
    ws.freeze_panes=f"A{FB}"


# ---------- Onlines Data (assessor pastes the export here) ----------
def _build_consultant_paste(wb):
    """Freeform paste sheet for consultants: headers in row 2, data from row 3.
    Onlines Data pulls every column by header name, so column order never matters -
    the consultant only has to keep the header NAMES matching Onlines Data row 3."""
    ws=wb.create_sheet(CONSULT_SHEET)
    ws["A1"]=("Paste the score extract here starting in cell A2 - group row, trait header row and "
              "all participants together, exactly as they come out of the extract. The group row "
              "lands in row 2, the trait headers in row 3 and the first participant in row 4. "
              "Column order does not matter, extra columns are ignored, and the identity names "
              "(Project Name, First Name, Last Name, Email, Completed Date, Status) are picked up "
              "from whichever of the two header rows they land in.")
    ws["A1"].font=F_IT; ws.merge_cells("A1:L1"); ws["A1"].alignment=A_WV
    ws.row_dimensions[1].height=42
    for r in range(CP_GROUP_ROW,CP_FIRST_DATA_ROW+2):
        for c in range(1,13):
            ws.cell(r,c).fill=C_PASTE
    ws.freeze_panes=f"A{CP_FIRST_DATA_ROW}"
    for c in ["A","B","C","D","E","F","G","H"]: ws.column_dimensions[c].width=16
    return ws

def _build_onlines_data(wb,sjt_variant=DEFAULT_SJT_VARIANT,selected_online=None):
    ws=wb.create_sheet(OD_SHEET)
    sv=SJT_VARIANTS[sjt_variant]
    all_groups={"MQ":("MQ.M5 Profile",MQ_DIMS),
                "OPQ":("OPQ32r",OPQ_DIMS+["Consistency Measure"]),
                "SJT":(sv["group"],sv["dims"]),
                "Verify":("Verify",["Inductive Percentile","Inductive Score","Deductive percentile",
                          "Deductive Score","Numerical Percentile","Numerical Score","Overall"])}
    order=["MQ","OPQ","SJT","Verify"]
    if selected_online is None: selected_online=order
    groups=[all_groups[t] for t in order if t in selected_online]
    ident=list(OD_IDENTITY)
    for i,h in enumerate(ident):
        col=openpyxl.utils.get_column_letter(i+1)
        ws.merge_cells(f"{col}1:{col}2")
        c=ws[f"{col}1"]; c.value=h; c.font=F_HDR; c.fill=C_HDR; c.alignment=A_WC
        ws[f"{col}3"]=h; ws[f"{col}3"].font=F_B; ws[f"{col}3"].fill=C_DIMBG
        ws[f"{col}3"].alignment=A_WV
        for r in (1,2,3): ws[f"{col}{r}"].border=BD
    col_idx=len(ident)+1          # traits start right after the identity columns
    for gname,dims in groups:
        gstart=col_idx
        for d in dims:
            col=openpyxl.utils.get_column_letter(col_idx)
            ws[f"{col}2"]=raw_header(d)                    # exact export header (lookup key 1)
            ws[f"{col}3"]=d                                 # clean trait name (lookup key)
            ws[f"{col}2"].font=F_IT; ws[f"{col}3"].font=F_B
            for r in (2,3):
                ws[f"{col}{r}"].border=BD; ws[f"{col}{r}"].alignment=A_WV
            ws[f"{col}3"].fill=C_DIMBG
            ws.column_dimensions[col].width=16
            col_idx+=1
        gend=col_idx-1
        c1=openpyxl.utils.get_column_letter(gstart); c2=openpyxl.utils.get_column_letter(gend)
        if gend>gstart: ws.merge_cells(f"{c1}1:{c2}1")
        h=ws[f"{c1}1"]; h.value=gname; h.font=F_HDR; h.fill=C_HDR; h.alignment=A_WC
        for cc in range(gstart,gend+1):
            ws[f"{openpyxl.utils.get_column_letter(cc)}1"].border=BD
    ws.freeze_panes=f"F{OD_FIRST_DATA_ROW}"
    # auto-pull every column from the Consultant Paste sheet by header name:
    # hidden helper row computes each column's MATCH once; data rows are cheap INDEX pulls
    cp_head=f"'{CONSULT_SHEET}'!$A${CP_HEADER_ROW}:${CP_LAST_COL}${CP_HEADER_ROW}"
    cp_grp=f"'{CONSULT_SHEET}'!$A${CP_GROUP_ROW}:${CP_LAST_COL}${CP_GROUP_ROW}"
    # row 2 holds the exact export header, row 3 the clean trait name — try both
    cp_data=f"'{CONSULT_SHEET}'!$A${CP_FIRST_DATA_ROW}:${CP_LAST_COL}${CP_FIRST_DATA_ROW+(OD_LAST_DATA_ROW-OD_FIRST_DATA_ROW)}"
    headers=ident+[ws.cell(3,c).value for c in range(len(ident)+1,col_idx)]
    for cidx,h in enumerate(headers,start=1):
        col=openpyxl.utils.get_column_letter(cidx)
        hcell=ws.cell(OD_HELPER_ROW,cidx)
        raw=str(ws.cell(2,cidx).value or h).replace('"','""')
        hh=str(h).replace('"','""')
        if cidx<=len(ident):
            # identity names land in the group row when the export merges them vertically,
            # but in the trait row when it doesn't — so search both
            hcell.value=(f'=IFERROR(MATCH("{hh}",{cp_head},0),'
                         f'IFERROR(MATCH("{hh}",{cp_grp},0),""))')
        else:
            # traits only ever live in the trait header row. Deliberately NOT falling back to
            # the group row: "Status" is both an identity name and an MQ trait, and a fallback
            # there would pull the identity value ("Completed") into a sten.
            hcell.value=(f'=IFERROR(MATCH("{raw}",{cp_head},0),'
                         f'IFERROR(MATCH("{hh}",{cp_head},0),""))')
        hcell.font=F_IT
        for r in range(OD_FIRST_DATA_ROW,OD_LAST_DATA_ROW+1):
            cp_r=r-OD_FIRST_DATA_ROW+1
            _ix=f'INDEX({cp_data},{cp_r},{col}${OD_HELPER_ROW})'
            cell=ws.cell(r,cidx)
            cell.value=f'=IF({col}${OD_HELPER_ROW}="","",IFERROR(IF({_ix}="","",{_ix}),""))'
            cell.fill=C_LOCK
    ws.row_dimensions[OD_HELPER_ROW].hidden=True
    for c,w in zip(["A","B","C","D","E","F"],[18,14,14,30,16,12]):
        ws.column_dimensions[c].width=w
    return ws

# ---------- T Chart ----------
TCHART_GUIDELINES=("Guidelines:\n"
"1. To be written in Second Person 'He/She'.\n"
"2. Mention strengths of the person (avoid replication of the behavioral indicators or OPQ statements)\n"
"3. Highlight the area of development such that it is (a) specific (b) actionable (c) constructive such that "
"it does not sound prescriptive (avoid 'should') and inspires the participant\n"
"4. Record critical evidences for these areas to justify being strength/areas of development and write them below.\n"
"5. For each development area, write two actionable development tips.\n"
"6. Appropriate linkage to the current or future role (if necessary)\n"
"7. Correctness/ accuracy (grammar, sentence construction, spell check)\n"
"8. 1 Strength area atleast (only if Rating 3-5) and 1 Development Area at least")

def _build_tchart(wb,framework,meta=None):
    """T Chart with live checks against the competency ratings in Detailed Integration.

    The sheet still never writes to Detailed Integration - it only reads the competency
    rating so it can warn when a low-rated competency is put down as a strength, or a
    high-rated one as a development area.
    """
    from openpyxl.formatting.rule import FormulaRule

    ws=wb.create_sheet("T Chart")
    comps=[]
    for c,_ in framework:
        if c not in comps: comps.append(c)

    # Detailed Integration geometry: data starts at row 8, one row per behaviour.
    # MATCH on the competency name lands on the FIRST row of that competency's block,
    # which is the merged anchor holding its rating.
    has_theme,has_level,_,_=_meta_cols(meta)
    n_pre=(1 if has_theme else 0)+1+(1 if has_level else 0)
    L=openpyxl.utils.get_column_letter
    comp_c=L(2) if has_theme else L(1)
    rate_c=L(n_pre+3)                      # behaviour, behaviour rating, competency rating
    di_first, di_last = 8, 8 + len(framework) - 1
    DI_NAME=f"'Detailed Integration'!${comp_c}${di_first}:${comp_c}${di_last}"
    DI_RATE=f"'Detailed Integration'!${rate_c}${di_first}:${rate_c}${di_last}"

    # guidelines block
    ws.merge_cells("B2:D12")
    g=ws["B2"]; g.value=TCHART_GUIDELINES; g.font=F_N; g.fill=C_LEG
    g.alignment=Alignment(wrap_text=True,vertical="top")
    for row in ws["B2:D12"]:
        for c in row: c.border=BD

    # summary block
    ws.merge_cells("B15:D15"); t=ws["B15"]; t.value="T Chart Summary"; t.font=F_HDR; t.fill=C_HDR; t.alignment=A_WC
    ws.merge_cells("B16:D16"); s=ws["B16"]; s.value="Strength Areas ( Min. 3 rating)"; s.font=F_B; s.fill=C_SUB; s.alignment=A_WV
    ws.merge_cells("B23:D23"); d=ws["B23"]; d.value="Development Areas (at least 1 competency)"; d.font=F_B; d.fill=C_SUB; d.alignment=A_WV

    rows=[(17,"Strength Area 1",True),(18,"Observation 1",False),(19,"Observation 2",False),
          (20,"Strength Area 2",True),(21,"Observation 1",False),(22,"Observation 2",False),
          (24,"Development Area 1",True),(25,"Observation 1",False),(26,"Observation 2",False),
          (27,"Development Area 2",True),(28,"Observation 1",False),(29,"Observation 2",False)]
    dd_rows=[17,20,24,27]
    strength_rows={17,20}

    # headers over the check columns
    for col,txt,w in (("E","Rating",12),("F","Check",56)):
        c=ws[f"{col}16"]; c.value=txt; c.font=F_HDR; c.fill=C_HDR; c.alignment=A_WC; c.border=BD

    for r,label,is_dd in rows:
        ws[f"B{r}"]=label; ws[f"B{r}"].font=F_B; ws[f"B{r}"].border=BD
        ws.merge_cells(f"C{r}:D{r}")
        cell=ws[f"C{r}"]; cell.value=""; cell.fill=C_IN
        cell.alignment=Alignment(wrap_text=True,vertical="top")
        for cc in ["C","D"]: ws[f"{cc}{r}"].border=BD
        if not is_dd:
            ws.row_dimensions[r].height=95
            continue

        # --- the competency's rating, read from Detailed Integration ---
        ws[f"E{r}"]=(f'=IF($C{r}="","",IFERROR(INDEX({DI_RATE},MATCH($C{r},{DI_NAME},0)),""))')
        ws[f"E{r}"].fill=C_LOCK; ws[f"E{r}"].font=F_B; ws[f"E{r}"].alignment=A_CTR
        ws[f"E{r}"].border=BD

        # --- duplicate: the same competency chosen in more than one area ---
        dup="+".join(f'($C${x}=$C{r})' for x in dd_rows)
        if r in strength_rows:
            wrong=(f'IF($E{r}<3,"Rated "&$E{r}&" - are you sure this is an Area of Strength?","")')
        else:
            wrong=(f'IF($E{r}>=3,"Rated "&$E{r}&" - are you sure this is an Area of Development?","")')
        ws[f"F{r}"]=(f'=IF($C{r}="","",'
                     f'IF(({dup})>1,"Already chosen in another area. ","")'
                     f'&IF($E{r}="","Competency rating is blank - check Detailed Integration.",'
                     f'IF({wrong}="","Good to go.",{wrong})))')
        ws[f"F{r}"].fill=C_LOCK; ws[f"F{r}"].font=F_N
        ws[f"F{r}"].alignment=Alignment(wrap_text=True,vertical="center")
        ws[f"F{r}"].border=BD

        # --- hidden status: 2 = conflict, 1 = rating not filled in yet, 0 = fine ---
        if r in strength_rows:
            bad=f'$E{r}<3'
        else:
            bad=f'$E{r}>=3'
        ws[f"G{r}"]=(f'=IF($C{r}="",0,IF(({dup})>1,2,IF($E{r}="",1,IF({bad},2,0))))')

    ws.column_dimensions["G"].hidden=True

    # --- competency ratings reference table, which also feeds the dropdowns ---
    ws["H6"]="Competency"; ws["I6"]="Rating"
    for cc in ("H","I"):
        ws[f"{cc}6"].font=F_HDR; ws[f"{cc}6"].fill=C_HDR
        ws[f"{cc}6"].alignment=A_WC; ws[f"{cc}6"].border=BD
    for i,comp in enumerate(comps):
        rr=FB+i
        ws[f"H{rr}"]=comp; ws[f"H{rr}"].font=F_N; ws[f"H{rr}"].border=BD; ws[f"H{rr}"].fill=C_DIMBG
        ws[f"I{rr}"]=(f'=IFERROR(INDEX({DI_RATE},MATCH($H{rr},{DI_NAME},0)),"")')
        ws[f"I{rr}"].font=F_B; ws[f"I{rr}"].fill=C_LOCK
        ws[f"I{rr}"].alignment=A_CTR; ws[f"I{rr}"].border=BD

    dv=DataValidation(type="list",formula1=f"=$H${FB}:$H${FB+len(comps)-1}",allow_blank=True)
    ws.add_data_validation(dv)
    for r in dd_rows: dv.add(f"C{r}")

    # --- conditional formatting driven by the hidden status column ---
    red=PatternFill("solid",fgColor="FFC7CE")      # a conflict
    amber=PatternFill("solid",fgColor="FFEB9C")    # rating not filled in yet
    for r in dd_rows:
        ws.conditional_formatting.add(
            f"C{r}:F{r}", FormulaRule(formula=[f"$G{r}=2"], fill=red, stopIfTrue=False))
        ws.conditional_formatting.add(
            f"C{r}:F{r}", FormulaRule(formula=[f"$G{r}=1"], fill=amber, stopIfTrue=False))

    for c,w in zip(["A","B","C","D","E","F","G","H","I"],[6,24,38,60,12,56,3,30,12]):
        ws.column_dimensions[c].width=w
    return ws

# ---------- Detailed Integration ----------
def _build_di(wb,selected,framework,tool_comps,mq_mode="5-point",online_paste=True,
              meta=None,show_summary=False,sjt_variant=DEFAULT_SJT_VARIANT):
    di=wb.create_sheet("Detailed Integration")
    di["A1"]="DETAILED INTEGRATION"; di["A1"].font=F_TITLE
    od_email=f"'{OD_SHEET}'!${OD_EMAIL_COL}${OD_FIRST_DATA_ROW}:${OD_EMAIL_COL}${OD_LAST_DATA_ROW}"
    od_first=f"'{OD_SHEET}'!${OD_FIRST_COL}${OD_FIRST_DATA_ROW}:${OD_FIRST_COL}${OD_LAST_DATA_ROW}"
    od_last=f"'{OD_SHEET}'!${OD_LAST_COL}${OD_FIRST_DATA_ROW}:${OD_LAST_COL}${OD_LAST_DATA_ROW}"
    tool_cols=[t for t in CANONICAL if t in selected]   # online-first ordering
    has_online=any(TOOLS[t]["family"]=="online" for t in tool_cols)
    # participant block. With the online paste sheets the email is a dropdown and the name
    # is looked up; without them there is nothing to look up, so both are typed.
    labels=([("Participant Email",2),("Participant Name",3)] if (has_online and online_paste)
            else [("Participant Name",2),("Participant Email",3)])
    for lab,r in labels+[("Assessor Name",4),("Date of Scoring",5)]:
        di[f"A{r}"]=lab; di[f"A{r}"].font=F_B; di[f"B{r}"]=""; di[f"B{r}"].border=BD
    if has_online and online_paste:
        di["B2"].fill=C_IN
        dv_email=DataValidation(type="list",formula1=f"={od_email}",allow_blank=True)
        di.add_data_validation(dv_email); dv_email.add("B2")
        di["B3"]=(f'=IFERROR(TRIM(INDEX({od_first},MATCH($B$2,{od_email},0))&" "&'
                  f'INDEX({od_last},MATCH($B$2,{od_email},0))),"")')
        di["B3"].fill=C_LOCK
    else:
        di["B2"].fill=C_IN; di["B3"].fill=C_IN
    di["B4"].fill=C_IN; di["B5"].fill=C_IN
    H=7                                   # header row (labels use 2-5, tags row 6)
    has_theme,has_level,theme_lbl,level_lbl=_meta_cols(meta)
    fixed=(([theme_lbl] if has_theme else [])+["Competency"]
           +([level_lbl] if has_level else [])
           +["Behaviour","Behaviour Rating","Competency Rating","Behaviour Avg","Check"])
    TOOL_OFF=len(fixed)                   # tool columns start here
    heads=fixed+tool_cols
    cols=[openpyxl.utils.get_column_letter(i+1) for i in range(len(heads))]
    for c,h in zip(cols,heads): di[f"{c}{H}"]=h
    _hdr(di,H,cols)
    # Consistency Measure (OPQ validity indicator) directly above the OPQ heading
    if "OPQ" in tool_cols and online_paste:
        oc=cols[TOOL_OFF+tool_cols.index("OPQ")]
        od_data=f"'{OD_SHEET}'!${OD_TRAIT_FIRST_COL}${OD_FIRST_DATA_ROW}:${OD_TRAIT_LAST_COL}${OD_LAST_DATA_ROW}"
        od_names=f"'{OD_SHEET}'!${OD_TRAIT_FIRST_COL}$3:${OD_TRAIT_LAST_COL}$3"
        cell=di[f"{oc}{H-1}"]
        cell.value=(f'=IF($B$2="","","Consistency: "&IFERROR(INDEX({od_data},'
                    f'MATCH($B$2,{od_email},0),MATCH("Consistency Measure",{od_names},0)),"-"))')
        cell.font=F_TAG; cell.alignment=A_CTR
    start=H+1; di_row=start; groups=[]
    n_pre=(1 if has_theme else 0)+1+(1 if has_level else 0)   # theme? comp level? -> behaviour
    theme_col=cols[0] if has_theme else None
    comp_col=cols[1] if has_theme else cols[0]
    level_col=cols[n_pre-1] if has_level else None
    beh_name_col=cols[n_pre]
    beh_col=cols[n_pre+1]; comp_rating_col=cols[n_pre+2]
    avg_col=cols[n_pre+3]; chk_col=cols[n_pre+4]
    status_col=openpyxl.utils.get_column_letter(len(cols)+1)   # hidden, drives the colours
    # per-tool row maps built from each tool's FILTERED framework
    tool_rows={}
    for t in tool_cols:
        filt=[(c,b) for c,b in framework if c in tool_comps[t]]
        if TOOLS[t]["family"]=="online":
            tool_rows[t]=({comp:gs for comp,behs,gs,ge in
                           _block_geometry(filt,_tool_slots(t,sjt_variant=sjt_variant))}
                          if filt else {})
        else:
            tool_rows[t]={(c,b):FB+i for i,(c,b) in enumerate(filt)}
    def _black(dcol,gs,ge):
        for rr in range(gs,ge+1):
            cell=di[f"{dcol}{rr}"]; cell.fill=C_BLACK; cell.border=BD
        if ge>gs: di.merge_cells(f"{dcol}{gs}:{dcol}{ge}")
    for comp,items in groupby(framework,key=lambda x:x[0]):
        items=list(items); gs=di_row; lvl_vals=[]
        for (c,beh) in items:
            di[f"{comp_col}{di_row}"]=comp; di[f"{beh_name_col}{di_row}"]=beh
            if has_level:
                lv=(meta.get("levels") or {}).get((comp,beh),"")
                di[f"{level_col}{di_row}"]=lv; lvl_vals.append(lv)
            di[f"{beh_col}{di_row}"]=""; di[f"{beh_col}{di_row}"].fill=C_IN
            for k,t in enumerate(tool_cols):
                if TOOLS[t]["family"]=="online" or comp not in tool_comps[t]:
                    continue  # online + unmapped handled at group level
                dcol=cols[TOOL_OFF+k]; sc=_score_col(t,meta); trow=tool_rows[t][(comp,beh)]
                di[f"{dcol}{di_row}"]=f'=IF({_ref(t)}!{sc}{trow}="","",{_ref(t)}!{sc}{trow})'
                di[f"{dcol}{di_row}"].fill=C_LOCK
            _cellsetup(di,di_row,cols)
            for cc in cols[n_pre+1:]: di[f"{cc}{di_row}"].alignment=A_CTR
            di_row+=1
        ge=di_row-1
        for k,t in enumerate(tool_cols):
            dcol=cols[TOOL_OFF+k]
            if comp not in tool_comps[t]:
                _black(dcol,gs,ge)                # not mapped -> one black block
                continue
            if TOOLS[t]["family"]!="online":
                continue
            blk_start=tool_rows[t][comp]
            cell=di[f"{dcol}{gs}"]
            ac=_avg_col(t,mq_mode,meta)
            cell.value=f'=IF({_ref(t)}!{ac}{blk_start}="","",{_ref(t)}!{ac}{blk_start})'
            cell.fill=C_LOCK; cell.font=F_B; cell.alignment=A_CTR
            if ge>gs: di.merge_cells(f"{dcol}{gs}:{dcol}{ge}")
            for rr in range(gs,ge+1): di[f"{dcol}{rr}"].border=BD
        # --- competency rating is entered by hand, and checked against the behaviours ---
        g=di[f"{comp_rating_col}{gs}"]
        g.value=""
        g.fill=C_IN; g.font=F_B; g.alignment=A_CTR

        avg=di[f"{avg_col}{gs}"]
        avg.value=f'=IFERROR(ROUND(AVERAGE({beh_col}{gs}:{beh_col}{ge}),0),"")'
        avg.fill=C_LOCK; avg.font=F_B; avg.alignment=A_CTR

        R=f"${comp_rating_col}${gs}"; A=f"${avg_col}${gs}"
        chk=di[f"{chk_col}{gs}"]
        chk.value=(f'=IF(AND({R}="",{A}=""),"",'
                   f'IF({R}="","Not entered yet - behaviour average is "&{A}&".",'
                   f'IF({A}="","No behaviour ratings entered yet.",'
                   f'IF({R}<>{A},"Behaviour average is "&{A}&" - you have entered "&{R}&".",'
                   f'"Good to go."))))')
        chk.fill=C_LOCK; chk.font=F_N
        chk.alignment=Alignment(wrap_text=True,vertical="center")

        st_=di[f"{status_col}{gs}"]
        st_.value=(f'=IF(AND({R}="",{A}=""),0,IF(OR({R}="",{A}=""),1,IF({R}<>{A},2,0)))')

        for cc in (comp_rating_col,avg_col,chk_col,status_col):
            if ge>gs: di.merge_cells(f"{cc}{gs}:{cc}{ge}")
            for rr in range(gs,ge+1): di[f"{cc}{rr}"].border=BD
        di.conditional_formatting.add(
            f"{comp_rating_col}{gs}:{chk_col}{ge}",
            FormulaRule(formula=[f"${status_col}${gs}=2"], fill=C_BAD, stopIfTrue=False))
        di.conditional_formatting.add(
            f"{comp_rating_col}{gs}:{chk_col}{ge}",
            FormulaRule(formula=[f"${status_col}${gs}=1"], fill=C_WARN, stopIfTrue=False))
        if has_theme:
            tv=di[f"{theme_col}{gs}"]
            tv.value=(meta.get("themes") or {}).get(comp,"")
            tv.font=F_B; tv.alignment=A_CTR
            if ge>gs: di.merge_cells(f"{theme_col}{gs}:{theme_col}{ge}")
            for rr in range(gs,ge+1): di[f"{theme_col}{rr}"].border=BD
        # competency written once per group; level merges only when it is one value
        _merge_down(di,comp_col,gs,ge,centre=False)
        if has_level: _merge_if_uniform(di,level_col,gs,ge,lvl_vals)
        groups.append((comp,gs,ge)); di_row=ge+1
    layout={"tool_cols":tool_cols,"cols":cols,"groups":groups,"start":start,
            "beh_col":beh_col,"comp_rating_col":comp_rating_col,"tool_off":TOOL_OFF,
            "comp_col":comp_col,"beh_name_col":beh_name_col,"theme_col":theme_col,
            "level_col":level_col,"meta":meta}
    di.column_dimensions[status_col].hidden=True
    widths=(([18] if has_theme else [])+[20]+([12] if has_level else [])
            +[44,15,16,14,44]+[10]*len(tool_cols))
    for c,w in zip(cols,widths): di.column_dimensions[c].width=w
    di.freeze_panes=f"A{start}"
    # summary table (optional)
    if not show_summary:
        return layout
    sr=di_row+2; di[f"A{sr}"]="COMPETENCY SUMMARY"; di[f"A{sr}"].font=F_B
    shd=sr+1
    for c,h in zip(["A","B","C","D","E"],["Competency","Individual","Cohort","Ideal","Previous DC"]): di[f"{c}{shd}"]=h
    _hdr(di,shd,["A","B","C","D","E"])
    for k,(comp,gs,ge) in enumerate(groups):
        rr=shd+1+k; di[f"A{rr}"]=comp
        di[f"B{rr}"]=f'=IF({comp_rating_col}{gs}="","",{comp_rating_col}{gs})'; di[f"B{rr}"].fill=C_LOCK
        for c in ["C","D","E"]: di[f"{c}{rr}"]=""; di[f"{c}{rr}"].fill=C_IN
        for c in ["A","B","C","D","E"]:
            di[f"{c}{rr}"].border=BD; di[f"{c}{rr}"].font=F_N
            if c!="A": di[f"{c}{rr}"].alignment=A_CTR

    return layout

# ---------- Integration (row format - for stack ranking) ----------
def _build_integration(wb,selected,framework,tool_comps,layout):
    """Transposed view: competencies and behaviours run across the top, each tool is a row.
    Cells for competencies a tool does not measure are blacked out, so a consultant can read
    across a single row to stack-rank participants on any tool."""
    L=openpyxl.utils.get_column_letter
    ws=wb.create_sheet("Integration")
    ws["A1"]="INTEGRATION  —  row view for stack ranking"; ws["A1"].font=F_TITLE
    DI="'Detailed Integration'"
    tool_cols=layout["tool_cols"]; dcols=layout["cols"]
    groups=layout["groups"]; beh_col=layout["beh_col"]; cr_col=layout["comp_rating_col"]
    meta=layout.get("meta"); has_theme,has_level,theme_lbl,level_lbl=_meta_cols(meta)
    comp_col_di=layout.get("comp_col","A"); beh_name_col_di=layout.get("beh_name_col","B")
    R_THEME=3 if has_theme else None
    R_COMP=(4 if has_theme else 3)
    R_CS=R_COMP+1; R_BEH=R_CS+1
    R_LVL=(R_BEH+1) if has_level else None
    R_BR=(R_LVL+1) if has_level else (R_BEH+1)
    labs=([(theme_lbl,R_THEME)] if has_theme else [])+[("Competency",R_COMP),
          ("Competency Score",R_CS),("Behaviour",R_BEH)]
    if has_level: labs.append((level_lbl,R_LVL))
    labs.append(("Behaviour Rating",R_BR))
    for lab,r in labs:
        c=ws[f"A{r}"]; c.value=lab; c.font=F_HDR; c.fill=C_HDR; c.alignment=A_WC; c.border=BD
    col=2
    for comp,gs,ge in groups:
        n=ge-gs+1; c1=L(col); c2=L(col+n-1)
        h=ws[f"{c1}{R_COMP}"]; h.value=comp; h.font=F_B; h.fill=C_SUB; h.alignment=A_WC
        s=ws[f"{c1}{R_CS}"]; s.value=f'=IF({DI}!{cr_col}{gs}="","",{DI}!{cr_col}{gs})'
        s.font=F_B; s.fill=C_LOCK; s.alignment=A_CTR
        if n>1:
            ws.merge_cells(f"{c1}{R_COMP}:{c2}{R_COMP}"); ws.merge_cells(f"{c1}{R_CS}:{c2}{R_CS}")
        if has_theme:
            th=ws[f"{c1}{R_THEME}"]
            th.value=(meta.get("themes") or {}).get(comp,"")
            th.font=F_B; th.fill=C_SUB; th.alignment=A_WC
            if n>1: ws.merge_cells(f"{c1}{R_THEME}:{c2}{R_THEME}")
            for i in range(n): ws[f"{L(col+i)}{R_THEME}"].border=BD
        for i in range(n):
            cc=L(col+i); dr=gs+i
            b=ws[f"{cc}{R_BEH}"]; b.value=f'={DI}!{beh_name_col_di}{dr}'
            b.font=F_N; b.alignment=A_WRAP; b.fill=C_DIMBG
            if has_level:
                lv=ws[f"{cc}{R_LVL}"]
                lv.value=(meta.get("levels") or {}).get((comp,None),"") or ""
                lv.font=F_N; lv.alignment=A_CTR; lv.fill=C_DIMBG
                if layout.get("level_col"):
                    lv.value=f'={DI}!{layout["level_col"]}{dr}'
            v=ws[f"{cc}{R_BR}"]; v.value=f'=IF({DI}!{beh_col}{dr}="","",{DI}!{beh_col}{dr})'
            v.font=F_N; v.alignment=A_CTR; v.fill=C_LOCK
            for r in [x for x in (R_THEME,R_COMP,R_CS,R_BEH,R_LVL,R_BR) if x]:
                ws[f"{cc}{r}"].border=BD
            ws.column_dimensions[cc].width=15
        col+=n
    r=R_BR+1
    for k,t in enumerate(tool_cols):
        lab=ws[f"A{r}"]; lab.value=t; lab.font=F_HDR; lab.fill=C_HDR; lab.alignment=A_WC; lab.border=BD
        tcol=dcols[layout.get("tool_off",6)+k]; col=2
        for comp,gs,ge in groups:
            n=ge-gs+1; c1=L(col); c2=L(col+n-1)
            if comp not in tool_comps[t]:
                for i in range(n):
                    cc=L(col+i); ws[f"{cc}{r}"].fill=C_BLACK; ws[f"{cc}{r}"].border=BD
                if n>1: ws.merge_cells(f"{c1}{r}:{c2}{r}")
            elif TOOLS[t]["family"]=="online":
                cell=ws[f"{c1}{r}"]
                cell.value=f'=IF({DI}!{tcol}{gs}="","",{DI}!{tcol}{gs})'
                cell.font=F_B; cell.fill=C_LOCK; cell.alignment=A_CTR
                for i in range(n): ws[f"{L(col+i)}{r}"].border=BD
                if n>1: ws.merge_cells(f"{c1}{r}:{c2}{r}")
            else:
                for i in range(n):
                    cc=L(col+i); dr=gs+i
                    cell=ws[f"{cc}{r}"]
                    cell.value=f'=IF({DI}!{tcol}{dr}="","",{DI}!{tcol}{dr})'
                    cell.font=F_N; cell.fill=C_LOCK; cell.alignment=A_CTR; cell.border=BD
            col+=n
        r+=1
    ws.column_dimensions["A"].width=22
    ws.freeze_panes="B3"
    return ws

# ---------- public API ----------
def _cluster_by_competency(framework):
    """Group all behaviours under the same competency together (first-appearance order),
    so the Detailed Integration competency average is always taken per competency, even
    if the uploaded framework interleaves competencies."""
    order, groups = [], {}
    for comp, beh in framework:
        if comp not in groups:
            groups[comp] = []; order.append(comp)
        groups[comp].append(beh)
    return [(c, b) for c in order for b in groups[c]]

def build_workbook(framework, selected_tools, sjt_variant=DEFAULT_SJT_VARIANT, tool_comps=None,
                   mq_mode="5-point", online_paste=True, meta=None, show_summary=False):
    """framework: list of (competency, behaviour). selected_tools: subset of CANONICAL.
    sjt_variant: "Managerial" or "Executive" (shapes the SJT sheet + Onlines Data columns).
    online_paste: True = the consultant pastes the online export, so the Consultant Paste and
    Onlines Data sheets are included and stens flow automatically. False = neither sheet is
    created and stens are typed straight into the online tool sheets.
    tool_comps: optional {tool: [competencies]} - a tool sheet only carries the competencies
    mapped to it; unmapped competencies show as a black block in Detailed Integration.
    Tools not in the dict get every competency."""
    if sjt_variant not in SJT_VARIANTS: sjt_variant=DEFAULT_SJT_VARIANT
    selected=[t for t in CANONICAL if t in selected_tools]
    framework=_cluster_by_competency(framework)   # ensures per-competency averaging
    all_comps=[]
    for c,_ in framework:
        if c not in all_comps: all_comps.append(c)
    # MQ in Low/High mode is a standalone profile: no competency mapping, absent from
    # Detailed Integration and Integration entirely (not even blacked-out cells).
    mq_standalone = ("MQ" in selected and mq_mode == "low-high")
    mapped = [t for t in selected if not (mq_standalone and t == "MQ")]
    tc={}
    for t in mapped:
        chosen=(tool_comps or {}).get(t)
        tc[t]=set(chosen) if chosen is not None else set(all_comps)
        if not any(c in tc[t] for c in all_comps):
            raise ValueError(f"'{t}' has no competencies mapped - untick the tool instead, "
                             f"or map at least one competency to it.")
    wb=openpyxl.Workbook(); wb.remove(wb.active)
    selected_online=[t for t in selected if TOOLS[t]["family"]=="online"]
    has_online=bool(selected_online)
    if has_online and online_paste:
        _build_consultant_paste(wb)                          # consultant pastes here (first sheet)
        _build_onlines_data(wb,sjt_variant,selected_online)  # auto-mapped mirror, selected tools only
    for name in selected:
        spec=TOOLS[name]
        if mq_standalone and name=="MQ":
            _build_mq_standalone(wb, MQ_DIMS, online_paste)
            continue
        fmeta=meta
        filt=[(c,b) for c,b in framework if c in tc[name]]
        if spec["kind"]=="online":
            dims=SJT_VARIANTS[sjt_variant]["dims"] if name=="SJT" else spec["dims"]
            _build_online(wb,name,dims,filt,mq_mode,online_paste,fmeta)
        elif spec["kind"]=="bars": _build_bars(wb,name,filt,fmeta)
        elif spec["kind"]=="bei":  _build_bei(wb,name,filt,fmeta)
    _build_tchart(wb,framework,meta)
    di_layout=_build_di(wb,mapped,framework,tc,mq_mode,online_paste,meta,show_summary,
                        sjt_variant)
    _build_integration(wb,mapped,framework,tc,di_layout)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines=False          # no gridlines
    return wb

def build_template():
    """Blank framework template.

    Four columns: Theme | Competency | Level | Behaviour. Only Competency and Behaviour are
    required - Theme and Level are optional, and if a client leaves them empty the generated
    workbook simply omits them. Whatever headings the client types are the headings used
    throughout the workbook, so 'Theme' can become 'Cluster' or anything else.
    """
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Competency Framework"
    ws.sheet_view.showGridLines=False
    ws["A1"]="COMPETENCY FRAMEWORK — upload template"; ws["A1"].font=F_TITLE
    ws["A2"]=("One row per behaviour. Repeat the competency on each of its behaviour rows, or "
              "leave it blank to inherit the one above.\n"
              "RENAME COLUMNS A AND C to whatever your client calls them - e.g. 'Themes' and "
              "'Competency Definition' - and your wording is used as the column heading right "
              "through the workbook. Both are optional: leave a column empty and it is left "
              "out everywhere.\n"
              "Keep the headings 'Competency' and 'Behaviour' exactly as they are - those two "
              "are how the file is read.\n"
              "Column A and Column B are written once per competency in the workbook, not "
              "repeated on every row. Column C is merged the same way when it holds one value "
              "per competency (a definition); if it varies behaviour to behaviour (a level), "
              "each row keeps its own value.")
    ws["A2"].fill=C_LEG; ws["A2"].font=F_IT; ws.merge_cells("A2:D2"); ws["A2"].alignment=A_WV
    ws.row_dimensions[2].height=92
    heads=["Theme","Competency","Level","Behaviour"]
    for i,h in enumerate(heads):
        c=ws.cell(4,i+1); c.value=h
        c.font=F_HDR; c.fill=C_HDR; c.alignment=Alignment(vertical="center"); c.border=BD
    ex=[("Leading Self","Drives Results","Core","Sets clear goals and holds the team accountable."),
        ("Leading Self","Drives Results","Core","Removes obstacles to keep work moving toward targets."),
        ("Leading Others","Leads People","Advanced","Coaches team members with regular, specific feedback.")]
    for i,row in enumerate(ex):
        for j,v in enumerate(row):
            cell=ws.cell(5+i,j+1); cell.value=v
            cell.fill=C_IN; cell.border=BD; cell.alignment=A_WRAP; cell.font=F_N
    for col,w in zip("ABCD",(24,28,16,64)): ws.column_dimensions[col].width=w
    ws.freeze_panes="A5"
    return wb


def parse_framework(source):
    """Read an uploaded framework.

    Competency and Behaviour are found by their header text wherever they sit. Any column
    to the LEFT of Competency is treated as the theme, and any column BETWEEN Competency and
    Behaviour as the level - both optional, and both keep whatever heading the client typed.

    Returns (framework, warnings, meta) where framework is [(competency, behaviour)] and meta
    carries {"theme_label","level_label","themes":{comp:theme},"levels":{(comp,beh):level}}.
    """
    wb = openpyxl.load_workbook(source, data_only=True)
    ws = wb["Competency Framework"] if "Competency Framework" in wb.sheetnames else wb.worksheets[0]

    def norm(v):
        return str(v).strip() if v is not None else ""

    # locate the header row and the two required columns
    header_row = comp_c = beh_c = None
    for r in range(1, min(ws.max_row, 30) + 1):
        found = {}
        for c in range(1, min(ws.max_column, 40) + 1):
            t = norm(ws.cell(r, c).value).lower()
            if t == "competency" and "comp" not in found:
                found["comp"] = c
            elif t in ("behaviour", "behavior") and "beh" not in found:
                found["beh"] = c
        if "comp" in found and "beh" in found and found["beh"] > found["comp"]:
            header_row, comp_c, beh_c = r, found["comp"], found["beh"]
            break
    if header_row is None:                     # no header - fall back to the old A/B layout
        header_row, comp_c, beh_c = 0, 1, 2

    theme_c = comp_c - 1 if comp_c > 1 else None
    level_c = comp_c + 1 if beh_c - comp_c > 1 else None
    theme_label = norm(ws.cell(header_row, theme_c).value) if (theme_c and header_row) else ""
    level_label = norm(ws.cell(header_row, level_c).value) if (level_c and header_row) else ""

    framework, warnings = [], []
    themes, levels = {}, {}
    last_comp = last_theme = ""
    for r in range(header_row + 1, ws.max_row + 1):
        comp = norm(ws.cell(r, comp_c).value)
        beh = norm(ws.cell(r, beh_c).value)
        theme = norm(ws.cell(r, theme_c).value) if theme_c else ""
        level = norm(ws.cell(r, level_c).value) if level_c else ""
        if theme:
            last_theme = theme
        if comp:
            last_comp = comp
            if last_theme:
                themes.setdefault(comp, last_theme)
        if not beh:
            continue
        if not last_comp:
            warnings.append(f"Row {r}: behaviour with no competency above it — skipped.")
            continue
        if last_theme:
            themes.setdefault(last_comp, last_theme)
        if level:
            levels[(last_comp, beh)] = level
        framework.append((last_comp, beh))

    if not framework:
        raise ValueError("No competency/behaviour rows found. Check the file has Competency "
                         "and Behaviour columns with at least one filled row.")

    meta = {"theme_label": theme_label or "Theme", "level_label": level_label or "Level",
            "themes": themes if any(themes.values()) else {},
            "levels": levels}
    return framework, warnings, meta


def parse_framework_text(text):
    """
    Parse a pasted framework. One behaviour per line, 'Competency, Behaviour'
    (comma or tab separated). A blank competency inherits the one above.
    A line with no separator is treated as a competency heading if it is the
    first thing seen, otherwise as a behaviour under the current competency.
    Returns (framework, warnings).
    """
    framework, warnings = [], []
    last_comp = ""
    for i, raw in enumerate(text.splitlines(), start=1):
        line = raw.strip()
        if not line:
            continue
        if line.lower() in ("competency\tbehaviour", "competency,behaviour", "competency behaviour"):
            continue  # header line
        if "\t" in line:
            parts = line.split("\t", 1)
        elif "," in line:
            parts = line.split(",", 1)
        else:
            parts = [line]
        if len(parts) == 2:
            comp, beh = parts[0].strip(), parts[1].strip()
            if comp:
                last_comp = comp
            if not beh:
                continue
            if not last_comp:
                warnings.append(f"Line {i}: behaviour with no competency yet — skipped.")
                continue
            framework.append((last_comp, beh))
        else:
            token = parts[0].strip()
            if not last_comp:
                last_comp = token          # first bare line = competency heading
            else:
                framework.append((last_comp, token))  # subsequent bare lines = behaviours
    if not framework:
        raise ValueError("Couldn't read any behaviours. Use one behaviour per line as "
                         "'Competency, Behaviour' (comma or tab between the two).")
    # typed frameworks carry no theme or level
    return framework, warnings, {"theme_label": "Theme", "level_label": "Level",
                                 "themes": {}, "levels": {}}


if __name__ == "__main__":
    # Running this file directly produces two example files so you can see it work.
    demo = [
        ("Drives Results", "Sets clear goals and holds the team accountable for outcomes."),
        ("Drives Results", "Removes obstacles and reprioritises to keep work moving toward targets."),
        ("Builds Collaboration", "Works across teams to solve shared problems and align on decisions."),
        ("Leads People", "Coaches team members with regular, specific feedback."),
    ]
    build_template().save("Framework_Template.xlsx")
    build_workbook(demo, ["OPQ", "MQ", "SJT", "Case Study",
                          "Business Role Play", "BEI"]).save("Sample_Scoresheet.xlsx")
    print("Wrote Framework_Template.xlsx and Sample_Scoresheet.xlsx")
